import requests
from bs4 import BeautifulSoup
import openpyxl

headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.135 Safari/537.36',
    }
console_category = '2'
host = "https://www.game29.ru"
product_page_url = "https://www.game29.ru/products"


def get_data_from_page(page_url, params, ws, current_excel_row_number):
    response = requests.get(url=page_url, headers=headers, params=params)
    print(response.url)
    page_soup = BeautifulSoup(response.content, "lxml")

    game_console_table = page_soup.find("div",
                                        class_=["col-lg-12", "col-md-12", "col-sm-12", "col-xs-12"],
                                        style="background: white;")
    game_console_rows = game_console_table.find_all("div", class_="row")
    print(len(game_console_rows))

    if len(game_console_rows) == 0:
        return False, current_excel_row_number

    for i in range(len(game_console_rows)):
        game_console_row = game_console_rows[i]

        # game_console_img_src = game_console_row.find("img").get("src")
        game_console_title = game_console_row.find_all("a")[1].text
        game_console_price = game_console_row.find_all("span")[1].text

        ws[f"A{current_excel_row_number}"] = game_console_title
        ws[f"B{current_excel_row_number}"] = game_console_price

        current_excel_row_number += 1

    return True, current_excel_row_number


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.worksheets[0])
    ws = wb.create_sheet("Игровые консоли")

    current_page_number = 1
    current_excel_row_number = 1
    params = {'category': console_category,
              'page': str(current_page_number)}
    page_has_rows, current_excel_row_number = get_data_from_page(product_page_url,
                                                                 params,
                                                                 ws,
                                                                 current_excel_row_number)

    while page_has_rows:
        current_page_number += 1
        params['page'] = str(current_page_number)
        page_has_rows, current_excel_row_number = get_data_from_page(product_page_url,
                                                                     params,
                                                                     ws,
                                                                     current_excel_row_number)

    wb.save("output.xlsx")


if __name__ == '__main__':
    main()
# 1 h 30 min